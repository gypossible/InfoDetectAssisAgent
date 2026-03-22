from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

logger = logging.getLogger(__name__)


class ExcelReaderError(Exception):
    """Excel 读取异常。"""


class ExcelWatchlistReader:
    HEADER_CANDIDATES = {
        "主体名称",
        "主体",
        "监测主体",
        "公司名称",
        "名称",
        "name",
    }

    def __init__(self, input_path: Path, target_column_letter: str = "B") -> None:
        self.input_path = input_path
        self.target_column_letter = target_column_letter.upper()

    def read_entities(self) -> list[str]:
        workbook_paths = self._resolve_workbooks()
        entities: list[str] = []
        try:
            for workbook_path in workbook_paths:
                entities.extend(self._read_entities_from_workbook(workbook_path))
        except Exception as exc:
            raise ExcelReaderError(f"读取 Excel 失败：{exc}") from exc

        entities = [item for item in entities if item]
        unique_entities = list(dict.fromkeys(entities))

        if not unique_entities:
            raise ExcelReaderError("未在 Excel 的 B 列中读取到有效主体名称，请检查文件内容。")

        logger.info(
            "已从 %s 个 Excel 文件中读取到 %s 个待监测主体。",
            len(workbook_paths),
            len(unique_entities),
        )
        return unique_entities

    def _resolve_workbooks(self) -> list[Path]:
        if not self.input_path.exists():
            raise FileNotFoundError(f"监测名单路径不存在：{self.input_path}")

        if self.input_path.is_file():
            return [self.input_path]

        workbook_paths = sorted(
            [
                path
                for path in self.input_path.iterdir()
                if path.is_file() and path.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}
            ]
        )
        if not workbook_paths:
            raise ExcelReaderError(f"目录中未找到可读取的 Excel 文件：{self.input_path}")
        return workbook_paths

    def _read_entities_from_workbook(self, workbook_path: Path) -> list[str]:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        entities: list[str] = []
        try:
            for worksheet in workbook.worksheets:
                for row in worksheet.iter_rows(
                    min_col=self._column_index,
                    max_col=self._column_index,
                    values_only=True,
                ):
                    value = row[0]
                    entity = self._normalize_entity(value)
                    if entity:
                        entities.append(entity)
        finally:
            workbook.close()

        logger.info("Excel 文件读取完成：%s，共提取 %s 条候选主体。", workbook_path.name, len(entities))
        return entities

    @property
    def _column_index(self) -> int:
        return column_index_from_string(self.target_column_letter)

    def _normalize_entity(self, value: object) -> str:
        if value is None:
            return ""
        entity = str(value).strip()
        if not entity:
            return ""
        if entity.lower() in self.HEADER_CANDIDATES:
            return ""
        return entity
