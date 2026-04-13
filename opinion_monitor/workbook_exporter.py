from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import column_index_from_string, get_column_letter

from .excel_reader import ExcelWatchlistReader


class AnnotatedWorkbookExporter:
    def __init__(self, target_column_letter: str = "B", news_limit: int = 10) -> None:
        self.target_column_letter = target_column_letter.upper()
        self.news_limit = max(1, news_limit)
        self._target_column_index = column_index_from_string(self.target_column_letter)
        self._body_alignment = Alignment(vertical="top", wrap_text=True)
        self._header_font = Font(bold=True)

    def build_entity_news_map(self, dataframe: pd.DataFrame) -> dict[str, list[str]]:
        if dataframe.empty:
            return {}

        entity_news_map: dict[str, list[str]] = {}
        ordered_dataframe = dataframe.copy()
        if "published_at_dt" in ordered_dataframe.columns:
            ordered_dataframe = ordered_dataframe.sort_values(
                by=["published_at_dt", "entity_name"],
                ascending=[False, True],
                na_position="last",
            )

        for _, row in ordered_dataframe.iterrows():
            entity_name = str(row.get("entity_name", "")).strip()
            if not entity_name:
                continue
            bucket = entity_news_map.setdefault(entity_name, [])
            if len(bucket) >= self.news_limit:
                continue
            formatted_item = self._format_news_item(row)
            if formatted_item and formatted_item not in bucket:
                bucket.append(formatted_item)
        return entity_news_map

    def export(
        self,
        source_workbook_path: Path,
        output_dir: Path,
        run_time: datetime,
        entity_news_map: dict[str, list[str]],
    ) -> Path:
        macro_enabled = source_workbook_path.suffix.lower() in {".xlsm", ".xltm"}
        output_suffix = ".xlsm" if macro_enabled else ".xlsx"
        output_path = output_dir / f"舆情写回名单_{run_time.strftime('%Y%m%d')}{output_suffix}"

        workbook = load_workbook(source_workbook_path, keep_vba=macro_enabled)
        try:
            for worksheet in workbook.worksheets:
                self._annotate_sheet(worksheet, entity_news_map)
            workbook.save(output_path)
        finally:
            workbook.close()
        return output_path

    def _annotate_sheet(self, worksheet, entity_news_map: dict[str, list[str]]) -> None:
        insert_at = self._target_column_index + 1
        worksheet.insert_cols(insert_at, amount=self.news_limit)

        header_row = self._detect_header_row(worksheet)
        for offset in range(self.news_limit):
            cell = worksheet.cell(row=header_row, column=insert_at + offset)
            cell.value = f"近一年舆情{offset + 1}"
            cell.font = self._header_font
            cell.alignment = self._body_alignment
            worksheet.column_dimensions[get_column_letter(insert_at + offset)].width = 28

        for row_index in range(1, worksheet.max_row + 1):
            raw_value = worksheet.cell(row=row_index, column=self._target_column_index).value
            entity_name = ExcelWatchlistReader.normalize_entity_value(raw_value)
            if not entity_name:
                continue

            news_items = entity_news_map.get(entity_name, [])
            for offset in range(self.news_limit):
                cell = worksheet.cell(row=row_index, column=insert_at + offset)
                cell.value = news_items[offset] if offset < len(news_items) else ""
                cell.alignment = self._body_alignment

    def _detect_header_row(self, worksheet) -> int:
        for row_index in range(1, min(worksheet.max_row, 20) + 1):
            value = worksheet.cell(row=row_index, column=self._target_column_index).value
            normalized = str(value).replace("\u3000", " ").strip().casefold() if value is not None else ""
            if normalized in ExcelWatchlistReader.NORMALIZED_HEADER_CANDIDATES:
                return row_index
        return 1

    @staticmethod
    def _format_news_item(row: pd.Series) -> str:
        published_at = str(row.get("published_at", "")).strip()
        source = str(row.get("source", "")).strip()
        title = str(row.get("title", "")).strip()
        snippet = str(row.get("snippet", "")).strip()

        parts = [part for part in [AnnotatedWorkbookExporter._compact_date(published_at), source, title] if part]
        text = " | ".join(parts)
        if not title and snippet:
            text = " | ".join(part for part in [AnnotatedWorkbookExporter._compact_date(published_at), source, snippet] if part)
        elif snippet:
            snippet = " ".join(snippet.split())
            if snippet and snippet not in title:
                text = f"{text}\n{snippet[:120]}".strip()
        return text.strip()

    @staticmethod
    def _compact_date(value: str) -> str:
        text = value.strip()
        if not text:
            return ""
        parsed = pd.to_datetime(text, errors="coerce")
        if pd.isna(parsed):
            return text[:19]
        return parsed.strftime("%Y-%m-%d")
